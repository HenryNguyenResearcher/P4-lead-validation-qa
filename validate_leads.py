"""
Lead Validation & QA Pipeline
Simulates the data validation workflow used in LinkedIn research / fulfilment VA roles.

Input:  data/raw_leads.csv   — raw lead list from client
Output: output/validated_leads.xlsx — QA report ready for Google Sheets upload
"""

import pandas as pd
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FILE = "data/raw_leads.csv"
OUTPUT_FILE = "output/validated_leads.xlsx"

# ── U.S. senior executive titles (mapped from JD: "U.S. corporate structures") ──
EXECUTIVE_TITLES = {
    "c-suite": ["ceo", "cto", "cfo", "coo", "cmo", "cpo", "chro", "ciso"],
    "vp": ["vp", "vice president"],
    "director": ["director"],
    "head": ["head of"],
    "manager": ["manager"],
    "executive": ["account executive", "executive"],
}

LINKEDIN_PATTERN = re.compile(r"^https://linkedin\.com/in/[a-zA-Z0-9_%-]+/?$")
EMAIL_PATTERN = re.compile(r"^[\w\.\+\-]+@[\w\-]+\.[a-zA-Z]{2,}$")

US_STATES_ABBR = [
    "AL","AK","AZ","AR","CA","CO","CT","DE","FL","GA","HI","ID","IL","IN","IA",
    "KS","KY","LA","ME","MD","MA","MI","MN","MS","MO","MT","NE","NV","NH","NJ",
    "NM","NY","NC","ND","OH","OK","OR","PA","RI","SC","SD","TN","TX","UT","VT",
    "VA","WA","WV","WI","WY","DC",
]

def classify_title_level(title: str) -> str:
    if not title:
        return "Unknown"
    t = title.lower()
    for level, keywords in EXECUTIVE_TITLES.items():
        if any(k in t for k in keywords):
            return level.replace("-", " ").title()
    return "Individual Contributor / Other"

def validate_linkedin(url: str) -> tuple[bool, str]:
    if not url:
        return False, "Missing"
    if not LINKEDIN_PATTERN.match(url.strip()):
        return False, "Invalid format"
    return True, "OK"

def validate_email(email: str) -> tuple[bool, str]:
    if not email:
        return False, "Missing"
    if not EMAIL_PATTERN.match(email.strip()):
        return False, "Invalid format"
    return True, "OK"

def validate_location(location: str) -> tuple[bool, str]:
    if not location:
        return False, "Missing"
    parts = location.strip().split()
    state = parts[-1].upper() if parts else ""
    if state in US_STATES_ABBR:
        return True, "OK"
    return False, "State not recognised — verify manually"

def normalize_name(val: str) -> str:
    return val.strip().title() if val else ""

def normalize_text(val: str) -> str:
    return val.strip() if val else ""

def calculate_confidence(row: dict) -> int:
    """Score 0–100 based on how complete and valid a lead record is."""
    score = 0
    if row["First Name"]: score += 15
    if row["Last Name"]: score += 15
    if row["Job Title"]: score += 15
    if row["Company"]: score += 15
    if row["linkedin_valid"] == "OK": score += 25
    if row["email_valid"] == "OK": score += 10
    if row["location_valid"] == "OK": score += 5
    return score

def build_flags(row: dict) -> str:
    flags = []
    if not row["First Name"]: flags.append("First name missing")
    if not row["Last Name"]: flags.append("Last name missing")
    if not row["Job Title"]: flags.append("Job title missing")
    if not row["Company"]: flags.append("Company missing")
    if row["linkedin_valid"] != "OK": flags.append(f"LinkedIn: {row['linkedin_valid']}")
    if row["email_valid"] == "Missing" and row["linkedin_valid"] != "OK":
        flags.append("No contact method")
    if row["location_valid"] != "OK": flags.append(f"Location: {row['location_valid']}")
    return " | ".join(flags) if flags else ""

def determine_status(confidence: int, flags: str) -> str:
    if confidence >= 85 and not flags:
        return "✅ Ready"
    elif confidence >= 60:
        return "⚠️ Review"
    else:
        return "🚩 Flag for QA"

def main():
    df = pd.read_csv(INPUT_FILE, dtype=str).fillna("")

    records = []
    for _, row in df.iterrows():
        first = normalize_name(row.get("first_name", ""))
        last = normalize_name(row.get("last_name", ""))
        title = normalize_text(row.get("job_title", ""))
        company = normalize_text(row.get("company", ""))
        location = normalize_text(row.get("location", ""))
        linkedin = normalize_text(row.get("linkedin_url", ""))
        email = normalize_text(row.get("email", ""))
        raw_notes = normalize_text(row.get("notes", ""))

        li_valid, li_status = validate_linkedin(linkedin)
        em_valid, em_status = validate_email(email)
        loc_valid, loc_status = validate_location(location)

        rec = {
            "ID": row.get("id", ""),
            "First Name": first,
            "Last Name": last,
            "Job Title": title,
            "Title Level": classify_title_level(title),
            "Company": company,
            "Location": location,
            "LinkedIn URL": linkedin,
            "linkedin_valid": li_status,
            "Email": email,
            "email_valid": em_status,
            "location_valid": loc_status,
            "original_notes": raw_notes,
        }
        rec["Confidence Score"] = calculate_confidence(rec)
        rec["Flags"] = build_flags(rec)
        rec["QA Status"] = determine_status(rec["Confidence Score"], rec["Flags"])
        rec["Validated By"] = "Automated QA v1.0"
        rec["Validated At"] = datetime.now().strftime("%Y-%m-%d %H:%M")

        records.append(rec)

    output_cols = [
        "ID", "First Name", "Last Name", "Job Title", "Title Level",
        "Company", "Location", "LinkedIn URL", "Email",
        "Confidence Score", "QA Status", "Flags", "Validated By", "Validated At",
    ]
    result_df = pd.DataFrame(records)[output_cols]

    # ── Summary stats ──
    total = len(result_df)
    ready = (result_df["QA Status"] == "✅ Ready").sum()
    review = (result_df["QA Status"] == "⚠️ Review").sum()
    flagged = (result_df["QA Status"] == "🚩 Flag for QA").sum()
    avg_conf = result_df["Confidence Score"].mean()

    summary_df = pd.DataFrame([
        {"Metric": "Total Leads Processed", "Value": total},
        {"Metric": "Ready (≥85, no flags)", "Value": ready},
        {"Metric": "Needs Review (60–84)", "Value": review},
        {"Metric": "Flagged for QA (<60 or critical gap)", "Value": flagged},
        {"Metric": "Average Confidence Score", "Value": f"{avg_conf:.1f}/100"},
        {"Metric": "Data Accuracy Rate", "Value": f"{ready/total*100:.1f}%"},
    ])

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        result_df.to_excel(writer, sheet_name="Validated Leads", index=False)
        summary_df.to_excel(writer, sheet_name="QA Summary", index=False)

    _apply_styling(OUTPUT_FILE)

    print(f"\n{'='*55}")
    print(f"  Lead Validation & QA Pipeline - Complete")
    print(f"{'='*55}")
    print(f"  Total leads processed : {total}")
    print(f"  [OK] Ready            : {ready}  ({ready/total*100:.0f}%)")
    print(f"  [!]  Needs review     : {review}  ({review/total*100:.0f}%)")
    print(f"  [X]  Flagged for QA   : {flagged}  ({flagged/total*100:.0f}%)")
    print(f"  Avg confidence score  : {avg_conf:.1f}/100")
    print(f"  Output saved to       : {OUTPUT_FILE}")
    print(f"{'='*55}\n")

def _apply_styling(filepath: str):
    wb = load_workbook(filepath)

    # ── Validated Leads sheet ──
    ws = wb["Validated Leads"]

    header_fill = PatternFill("solid", fgColor="1F497D")
    ready_fill = PatternFill("solid", fgColor="C6EFCE")
    review_fill = PatternFill("solid", fgColor="FFEB9C")
    flag_fill = PatternFill("solid", fgColor="FFC7CE")
    alt_fill = PatternFill("solid", fgColor="F2F6FC")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_widths = {
        "A": 5, "B": 14, "C": 14, "D": 24, "E": 22,
        "F": 20, "G": 20, "H": 42, "I": 28,
        "J": 17, "K": 16, "L": 40, "M": 20, "N": 18,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = Font(bold=True, color="FFFFFF", size=10)
            else:
                status_val = ws.cell(row=row_idx, column=11).value or ""
                if "Ready" in status_val:
                    cell.fill = ready_fill
                elif "Review" in status_val:
                    cell.fill = review_fill
                elif "Flag" in status_val:
                    cell.fill = flag_fill
                elif row_idx % 2 == 0:
                    cell.fill = alt_fill
                cell.font = Font(size=10)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── QA Summary sheet ──
    ws2 = wb["QA Summary"]
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 20
    for row in ws2.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top")
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = Font(bold=True, color="FFFFFF", size=10)
            else:
                cell.font = Font(size=10)

    wb.save(filepath)

if __name__ == "__main__":
    main()
